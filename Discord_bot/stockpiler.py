import pymysql.cursors
import discord
from discord.ext.commands import Bot, CommandNotFound, bot_has_permissions
from multiprocessing import Process
import re
import pytz
from datetime import datetime, timedelta, timezone
import asyncio
from PIL import Image, ImageDraw, ImageFont
import random
import urllib.request, json
import time
import math
import os
import numbers


from openpyxl import load_workbook
from shutil import copyfile

intents = discord.Intents.default()
intents.members = True

running = False

bot = Bot(command_prefix='!', case_insensitive=True, intents=intents)
active_stockpiles = "0"

connection = pymysql.connect(host='www.stockpiler.net',
                             user='admin',
                             password='password',
                             db='database_name',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)

bot.remove_command('help')

server_time = 'US/Pacific'  # DB servers time zone
mytimezone = 'US/Eastern'  # For a list:  https://gist.github.com/JellyWX/913dfc8b63d45192ad6cb54c829324ee
stockalarm_lastran = time.time()

with open("beta/server/item_list_rev2.json", "r") as f:
    temp = json.loads(f.read())
item_list = {}
for g in temp:
    item_list[g["pcname"]] = g

dev_guild_name = "Stockpiler"  # If you have _dev in the file name, the bot will only operate for the specified discord.

debug = False
if "_dev" in os.path.basename(__file__):
    debug = True

admin_list = [185863100942123008]


def get_war_number():
    with urllib.request.urlopen(
            "https://war-service-live.foxholeservices.com/api/worldconquest/war") as url:
        data = json.loads(url.read())
    war_data = data
    if len(war_data) > 0:
        return war_data["warNumber"]

warNumber = get_war_number()

def check_war_change():
    try:
        with urllib.request.urlopen(
                "https://war-service-live.foxholeservices.com/api/worldconquest/war") as url:
            data = json.loads(url.read())
    except:
        print("Failed to load WAR date.")
        return
    war_data = data

    query = "SELECT * FROM system_vars WHERE id = 1"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            war_data_server = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return
    if len(war_data) > 0:
        EST = pytz.timezone(mytimezone)
        now = datetime.now(EST)
        war = war_data["warNumber"] + 1
        if war_data["winner"] != "None":
            if war_data_server["string"] != "1" and war_data_server["string"] != "3":
                future_plus_24_hours = datetime.now() + timedelta(hours=24)
                query = f"UPDATE system_vars SET `int` = '{war}' AND `datetime` = '{future_plus_24_hours}' AND `string` = '1' WHERE `id` = '1'"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query)
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
        elif war_data_server["string"] == "3":
            query = f"UPDATE system_vars SET `string` = '0' WHERE `id` = '1'"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    connection.commit()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
        else:
            future_minus_24_hours = now - timedelta(hours=24)
            if EST.localize(war_data_server["datetime"]) < future_minus_24_hours:
                # reset war system

                # discord_stockpiles table
                print(f"WAR {(war - 1)} has ended, executing changeover!")
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute("TRUNCATE TABLE discord_stockpiles")
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
                # stock table
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute("TRUNCATE TABLE stock_stock")
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
                # log table
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute("TRUNCATE TABLE stock_log")
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
                # users table
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute("TRUNCATE TABLE stock_users")
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
                # stockpiles_list table
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute("TRUNCATE TABLE stockpiles_list")
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return

                # recent_discord_messages table
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute("TRUNCATE TABLE recent_discord_messages")
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return

                # OVerwatch
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute("TRUNCATE TABLE overwatch_watch_list")
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return

                query = f"UPDATE system_vars SET `string` = '3' WHERE `id` = '1'"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query)
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return

                print(f"War changeover complete. War {war} started!")


async def display_current_stockpiles(ctx, args, discords_list_data={}):
    if len(discords_list_data) == 0:
        query = f"SELECT * FROM discords_list WHERE discord_id = %s"
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query, [ctx.guild.id])
                discords_list_data = cursor.fetchone()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return
    # -- Check for data validity -- #
    user_roles = ctx.author.roles
    is_admin = False
    for k in user_roles:
        if k.id == int(discords_list_data['admin_stockpiles_role']):
            is_admin = True
        for j in admin_list:
            if int(j) == k.id:
                is_admin = True
    if discords_list_data is None or discords_list_data['stockpile_code_channel'] is None:
        if ctx.author.guild_permissions.administrator or is_admin:
            await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup. Please run **!stock setup wizard**")
        else:
            await ctx.send(
                f"{ctx.message.author.mention} Stockpiler not correctly setup. Please have an ADMIN run **!stock setup wizard**")
        return
    if discords_list_data['stockpile_codes_message'] is None:
        await ctx.send(f"{ctx.message.author.mention} No stockpiles are setup for *{ctx.guild.name}*. Please run **!stock add** to add one.")
        return

    # -- Getting stockpile code data -- #
    codes_channel = bot.get_channel(int(discords_list_data['stockpile_code_channel']))
    if codes_channel is not None:
        try:
            stockpile_data = await codes_channel.fetch_message(int(discords_list_data['stockpile_codes_message']))
        except discord.NotFound as response:
            await ctx.send(
                f"{ctx.message.author.mention} Error, stockpile codes message missing!\n"
                f"Use the command !stock setup codeschannel #name-of-channel\n"
                f"Ex: !stock setup codeschannel #stockpile-codes\n"
                f"Then try adding again.")
            print(response)
            return

        data_message = stockpile_data.content[
                  162:]  # its the line breaks that make it harder to delete the warning, need some idea on removing it.
        if len(data_message) < 5:
            await ctx.send(
                f"{ctx.message.author.mention} No stockpiles have been added for this discord.\n"
                f"To add a new stockpile, the command is\n"
                f"!stock add code Stockpile Name\n"
                f"Ex: !stock add 135795 Blemish")
            return
        stockpiles = json.loads(data_message)
        if len(args) == 2 and len(str(args[1])) > 0:
            mention = args[1]
        else:
            mention = ctx.message.author.mention
        send_message_header = f"{mention}\n**Stockpiles for {ctx.guild.name}**\n"
        if discords_list_data['beta'] == 1:
            send_message_footer = "\nhttps://beta.stockpiler.net"
        else:
            send_message_footer = "\nhttps://www.stockpiler.net"

        send_message = ''
        for pile in stockpiles:
            send_message = send_message + f"**{pile['code']}** - {pile['name']}\n"

        if len(send_message) > 0:
            send_message = send_message_header + send_message + send_message_footer
            await codes_channel.send(send_message)
            return
        else:
            await ctx.send(
                f"{ctx.message.author.mention} Error loading stockpile message. Error Code: vrXstDXp")
            return
    else:
        if ctx.author.guild_permissions.administrator or is_admin:
            await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup or no stockpiles exist. Please run **!stock setup** or **!stock add**")
        else:
            await ctx.send(
                f"{ctx.message.author.mention} Stockpiler not correctly setup or no stockpiles exist. Please have an ADMIN run **!stock setup** or **!stock add**")
            return



async def add_stockpile(ctx, args):  # !stock add command
    banned = check_ban_status(ctx)
    if banned:
        return

    query = f"SELECT * FROM discords_list WHERE discord_name = %s"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query, [ctx.guild.name])
            discord_data = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return

    if len(discord_data) > 0:
        commands_channel = ctx
        if len(discord_data["commands_channel"]) > 0:
            commands_channel = get_command_channel(ctx)
        else:
            commands_channel = ctx

        if len(args) >= 3:
            code = args[1]
            counter = 0
            stockpile_name = ""
            for x in args:
                if counter < 2:  # Skips the word Add and the stockpile code
                    counter += 1
                    continue
                stockpile_name += x
                counter += 1
                if len(args) > counter:
                    stockpile_name += " "
            if len(stockpile_name) > 100:
                await commands_channel.send(
                    f"{ctx.message.author.mention} Stockpile name is limited to 100 Characters. Your name was {len(stockpile_name)}. Please pick a shorter name.")
                return

            user_roles = ctx.author.roles
            is_admin = False
            for k in user_roles:
                if k.id == int(discord_data['admin_stockpiles_role']):
                    is_admin = True
                for j in admin_list:
                    if int(j) == k.id:
                        is_admin = True



            if discord_data['stockpile_code_channel'] is None:
                if ctx.author.guild_permissions.administrator or is_admin:
                    await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup. Please run **!stock setup wizard**")
                else:
                    await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup. Please have an ADMIN run **!stock setup wizard**")
                    return
            code_warning = "DO NOT DELETE OR STOCKPILER WILL BREAK!!\n" \
                           "Stockpiler will edit this as needed to track your stockpile codes\n" \
                           "Stockpile codes are stored here and not on stockpiler.\n"
            await commands_channel.send(
                f"{ctx.message.author.mention} Adding Stockpile {stockpile_name}, please standby, this may take several minutes...")
            codes_channel = bot.get_channel(int(discord_data['stockpile_code_channel']))
            stockpile_data = None
            if codes_channel is not None:
                code_message_id = None
                if discord_data['stockpile_codes_message'] is None or len(discord_data['stockpile_codes_message']) == 0:
                    code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                    discord_data['stockpile_codes_message'] = str(code_message_id)
                else:
                    code_message_id = discord_data['stockpile_codes_message']
                if code_message_id is not None and len(str(code_message_id)) > 0:
                    try:
                        stockpile_data = await codes_channel.fetch_message(int(discord_data['stockpile_codes_message']))
                    except discord.NotFound as response:
                        code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                        discord_data['stockpile_codes_message'] = str(code_message_id)
                        try:
                            stockpile_data = await codes_channel.fetch_message(
                                int(discord_data['stockpile_codes_message']))
                        except discord.NotFound as response:
                            pass
                else:
                    code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                    discord_data['stockpile_codes_message'] = str(code_message_id)
                    try:
                        stockpile_data = await codes_channel.fetch_message(int(discord_data['stockpile_codes_message']))
                    except discord.NotFound as response:
                        code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                        discord_data['stockpile_codes_message'] = str(code_message_id)
                        try:
                            stockpile_data = await codes_channel.fetch_message(
                                int(discord_data['stockpile_codes_message']))
                        except discord.NotFound as response:
                            pass

                if code_message_id is None or len(str(code_message_id)) == 0 or stockpile_data is None:
                    if ctx.author.guild_permissions.administrator or is_admin:
                        await ctx.send(
                            f"{ctx.message.author.mention} Stockpiler not correctly setup. Please run **!stock setup wizard**")
                    else:
                        await ctx.send(
                            f"{ctx.message.author.mention} Stockpiler not correctly setup. Please have an ADMIN run **!stock setup wizard**")
                    return
                message = stockpile_data.content[
                          162:]  # its the line breaks that make it harder to delete the warning, need some idea on removing it.

                # Create a random ID to use as a ref in stockpiles database

                foundcode = False
                loop_count = 0
                while not foundcode:
                    if loop_count > 10:
                        # error relates to the bot generating 10 random codes but they are all aready assigned to stockples.
                        await get_command_channel(ctx, command_channel).send(
                            f"{ctx.message.author.mention} Error adding stockpile, please try again. Error Code: 4La9KmOQ")
                        return
                    random_number = random.randint(0, 99999999999999999999)
                    hex_number = str(hex(random_number))[2:]
                    query = f"SELECT `id` FROM stockpiles_list WHERE `idhash`='{hex_number}'"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query)
                            does_stockpile_exist = cursor.fetchone()
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
                    if does_stockpile_exist is None:
                        foundcode = True
                    loop_count += 1
                #stockpile Json data from discord.
                stockpile_data_json = {"code": code, "name": stockpile_name, "id": hex_number}
                stock_data_before = []
                if len(message) > 5:
                    stock_data = json.loads(message)
                    stock_data_before = json.loads(message)
                    stock_data.append(stockpile_data_json)
                else:
                    stock_data = [stockpile_data_json]
                # Check if stockpile code is already in use.
                if len(stock_data_before) > 0:
                    for g in stock_data_before:
                        if g['code'] == code:
                            await commands_channel.send(
                                f"{ctx.message.author.mention} Error adding stockpile, the stockpile code {code} is already in use with the **{g['name']}** stockpile.")
                            await display_current_stockpiles(ctx, args, discord_data)
                            return
                #insert stockpile data into database in the 'stockpiles_list' table
                query = f"INSERT INTO stockpiles_list (name, idhash, creator_discord_id, for_war_number, discordid) VALUES (%s, %s, %s, %s, %s)"
                data = (stockpile_name, hex_number, ctx.author.id, warNumber, discord_data['id'])
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query, data)
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    await get_command_channel(ctx, command_channel).send(
                        f"{ctx.message.author.mention} Error adding stockpile, please try again. Error Code: FMxxlOfE")
                    print(except_detail)
                    return

                stockpile_data_json = json.dumps(stock_data)
                await stockpile_data.edit(content=(code_warning + stockpile_data_json)) #add the new stockpile to the json data


                query = f"INSERT INTO stock_log (notes, user_id, log_type, stockpile_id) VALUES (%s, %s, %s, %s)"
                notes = f'<font color="red"><b>{ctx.author.name}</b></font> has created this stockpile this stockpile.'
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query, [notes, ctx.author.id, 'notice', hex_number])
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
                await display_current_stockpiles(ctx, args, discord_data)  # List all stockpiles with this new one.
                await check_updates(ctx, hex_number)
                await codes_channel.send(f"Stockpile {stockpile_name} added!")
                await commands_channel.send(f"{ctx.message.author.mention} Stockpile {stockpile_name} added!")
                # Try to delete the add command the person posted, we don't really care if it fails.
                try:
                    await ctx.message.delete()
                except:
                    return


            else:
                await commands_channel.send(
                    f"{ctx.message.author.mention} Error locating the Stockpile codes channel, check bots perms for the channel.\n"
                    f"If needed reconfigure the bot to the correct channel.\n"
                    f"!stock setup codeschannel #stockpile-codes")
                return


        else:
            await commands_channel.send(
                f"{ctx.message.author.mention} To add a new stockpile, the command is\n"
                f"!stock add code Stockpile Name\n"
                f"Ex: !stock add 135795 Blemish")
            return
    else:
        await ctx.send(
            f"{ctx.message.author.mention} This discord is not setup with my yet. Use the command **!stock setup wizard** to start the process!\n"
            "After im setup you can add stockpiles for me to keep track of here!")
        return


async def del_stockpile(ctx, args):  # !stock add command
    banned = check_ban_status(ctx)
    if banned:
        return



    query = f"SELECT * FROM discords_list WHERE discord_name = %s"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query, [ctx.guild.name])
            discord_data = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return

    if len(discord_data) > 0:
        commands_channel = ctx
        if len(discord_data["commands_channel"]) > 0:
            commands_channel = get_command_channel(ctx)
        else:
            commands_channel = ctx
        if len(args) <= 1:
            await commands_channel.send(
                f"{ctx.message.author.mention} To delete a stockpile, use the stockpile ID. **!stock del 123456** See the current stockpile list with **!stock codes**")
            return
        elif len(args) > 2:
            await commands_channel.send(
                f"{ctx.message.author.mention} You can only delete 1 stockpile at a time, the comand is **!stock del 123456** replacing 123456 with the stockpile code.")
            return

        user_roles = ctx.author.roles
        is_admin = False
        for k in user_roles:
            if k.id == int(discord_data['admin_stockpiles_role']):
                is_admin = True
            for j in admin_list:
                if int(j) == k.id:
                    is_admin = True



        if discord_data['stockpile_code_channel'] is None:
            if ctx.author.guild_permissions.administrator or is_admin:
                await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup. Please run **!stock setup wizard**")
            else:
                await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup. Please have an ADMIN run **!stock setup wizard**")
                return
        code_warning = "DO NOT DELETE OR STOCKPILER WILL BREAK!!\n" \
                       "Stockpiler will edit this as needed to track your stockpile codes\n" \
                       "Stockpile codes are stored here and not on stockpiler.\n"

        codes_channel = bot.get_channel(int(discord_data['stockpile_code_channel']))
        stockpile_data = None
        if codes_channel is not None:
            code_message_id = None
            if discord_data['stockpile_codes_message'] is None or len(discord_data['stockpile_codes_message']) == 0:
                code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                discord_data['stockpile_codes_message'] = str(code_message_id)
            else:
                code_message_id = discord_data['stockpile_codes_message']
            if code_message_id is not None and len(str(code_message_id)) > 0:
                try:
                    stockpile_data = await codes_channel.fetch_message(int(discord_data['stockpile_codes_message']))
                except discord.NotFound as response:
                    code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                    discord_data['stockpile_codes_message'] = str(code_message_id)
                    try:
                        stockpile_data = await codes_channel.fetch_message(
                            int(discord_data['stockpile_codes_message']))
                    except discord.NotFound as response:
                        pass
            else:
                code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                discord_data['stockpile_codes_message'] = str(code_message_id)
                try:
                    stockpile_data = await codes_channel.fetch_message(int(discord_data['stockpile_codes_message']))
                except discord.NotFound as response:
                    code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                    discord_data['stockpile_codes_message'] = str(code_message_id)
                    try:
                        stockpile_data = await codes_channel.fetch_message(
                            int(discord_data['stockpile_codes_message']))
                    except discord.NotFound as response:
                        pass

            if code_message_id is None or len(str(code_message_id)) == 0 or stockpile_data is None:
                if ctx.author.guild_permissions.administrator or is_admin:
                    await ctx.send(
                        f"{ctx.message.author.mention} Stockpiler not correctly setup. Please run **!stock setup wizard**")
                else:
                    await ctx.send(
                        f"{ctx.message.author.mention} Stockpiler not correctly setup. Please have an ADMIN run **!stock setup wizard**")
                return
            message = stockpile_data.content[
                      162:]  # its the line breaks that make it harder to delete the warning, need some idea on removing it.

            # Create a random ID to use as a ref in stockpiles database



            #stockpile Json data from discord.
            stock_data_before = []
            if len(message) > 5:
                stock_data = json.loads(message)
            else:
                await commands_channel.send(
                    f"{ctx.message.author.mention} No stockpiles exist, what are you trying to delete? Use command **!stock add** to add a stockpile.")
                await display_current_stockpiles(ctx, args, discord_data)
                return
            counter = 0
            message = ''
            stockpile_name = ''
            for x in stock_data:
                if x['code'] == args[1]:
                    query = f"SELECT * FROM stockpiles_list WHERE `idhash`='{x['id']}'"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query)
                            does_stockpile_exist = cursor.fetchone()
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
                    if does_stockpile_exist is not None:
                        query = f"UPDATE stockpiles_list SET `active` = %s WHERE `id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, ['0', does_stockpile_exist['id']])
                                connection.commit()
                                message += f"Stockpile **{x['name']}** has been deleted!\n"
                        except pymysql.err.ProgrammingError as except_detail:
                            message += "Error deleting stockpile in database, iv removed it from the discord but " \
                                       "cannot locate it in the database. This may or may not cause issues down the line. Error Code: CJKWt0Iy\n"
                            print(except_detail)
                    else:
                        message += "Error deleting stockpile in database, iv removed it from the discord but " \
                                   "cannot locate it in the database. This may or may not cause issues down the line. Error Code: K1QjAWIz\n"
                    stockpile_name = x['name']
                    del stock_data[counter]
                    await commands_channel.send(message)
                    await stockpile_data.edit(content=(code_warning + json.dumps(stock_data)))
                    await codes_channel.send(f"Stockpile {stockpile_name} DELETED!! by: {ctx.message.author.mention}")
                    await display_current_stockpiles(ctx, args, discord_data)

                counter += 1

            # Try to delete the add command the person posted, we don't really care if it fails.
            try:
                await ctx.message.delete()
            except:
                return

    else:
        await ctx.send(
            f"{ctx.message.author.mention} This discord is not setup with my yet. Use the command **!stock setup wizard** to start the process!\n"
            "After im setup you can add stockpiles for me to keep track of here!")
        return


async def rename_stockpile(ctx, args):  # !stock add command
    banned = check_ban_status(ctx)
    if banned:
        return



    query = f"SELECT * FROM discords_list WHERE discord_name = %s"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query, [ctx.guild.name])
            discord_data = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return

    if len(discord_data) > 0:
        commands_channel = ctx
        if len(discord_data["commands_channel"]) > 0:
            commands_channel = get_command_channel(ctx)
        else:
            commands_channel = ctx
        if len(args) <= 1:
            await commands_channel.send(
                f"{ctx.message.author.mention} To rename a stockpile, use the stockpile ID. **!stock rename 123456 new name** See the current stockpile list with **!stock codes**")
            return

        user_roles = ctx.author.roles
        is_admin = False
        for k in user_roles:
            if k.id == int(discord_data['admin_stockpiles_role']):
                is_admin = True
            for j in admin_list:
                if int(j) == k.id:
                    is_admin = True



        if discord_data['stockpile_code_channel'] is None:
            if ctx.author.guild_permissions.administrator or is_admin:
                await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup. Please run **!stock setup wizard**")
            else:
                await ctx.send(f"{ctx.message.author.mention} Stockpiler not correctly setup. Please have an ADMIN run **!stock setup wizard**")
                return
        code_warning = "DO NOT DELETE OR STOCKPILER WILL BREAK!!\n" \
                       "Stockpiler will edit this as needed to track your stockpile codes\n" \
                       "Stockpile codes are stored here and not on stockpiler.\n"

        codes_channel = bot.get_channel(int(discord_data['stockpile_code_channel']))
        stockpile_data = None
        if codes_channel is not None:
            code_message_id = None
            if discord_data['stockpile_codes_message'] is None or len(discord_data['stockpile_codes_message']) == 0:
                code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                discord_data['stockpile_codes_message'] = str(code_message_id)
            else:
                code_message_id = discord_data['stockpile_codes_message']
            if code_message_id is not None and len(str(code_message_id)) > 0:
                try:
                    stockpile_data = await codes_channel.fetch_message(int(discord_data['stockpile_codes_message']))
                except discord.NotFound as response:
                    code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                    discord_data['stockpile_codes_message'] = str(code_message_id)
                    try:
                        stockpile_data = await codes_channel.fetch_message(
                            int(discord_data['stockpile_codes_message']))
                    except discord.NotFound as response:
                        pass
            else:
                code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                discord_data['stockpile_codes_message'] = str(code_message_id)
                try:
                    stockpile_data = await codes_channel.fetch_message(int(discord_data['stockpile_codes_message']))
                except discord.NotFound as response:
                    code_message_id = await create_stockpile_database(codes_channel, code_warning, discord_data, ctx)
                    discord_data['stockpile_codes_message'] = str(code_message_id)
                    try:
                        stockpile_data = await codes_channel.fetch_message(
                            int(discord_data['stockpile_codes_message']))
                    except discord.NotFound as response:
                        pass

            if code_message_id is None or len(str(code_message_id)) == 0 or stockpile_data is None:
                if ctx.author.guild_permissions.administrator or is_admin:
                    await ctx.send(
                        f"{ctx.message.author.mention} Stockpiler not correctly setup. Please run **!stock setup wizard**")
                else:
                    await ctx.send(
                        f"{ctx.message.author.mention} Stockpiler not correctly setup. Please have an ADMIN run **!stock setup wizard**")
                return
            message = stockpile_data.content[
                      162:]  # its the line breaks that make it harder to delete the warning, need some idea on removing it.

            # Create a random ID to use as a ref in stockpiles database



            #stockpile Json data from discord.
            stock_data_before = []
            if len(message) > 5:
                stock_data = json.loads(message)
            else:
                await commands_channel.send(
                    f"{ctx.message.author.mention} No stockpiles exist, what are you trying to delete? Use command **!stock add** to add a stockpile.")
                await display_current_stockpiles(ctx, args, discord_data)
                return
            counter = 0
            message = ''
            stockpile_name = ''
            for x in stock_data:
                if x['code'] == args[1]:
                    query = f"SELECT * FROM stockpiles_list WHERE `idhash`='{x['id']}'"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query)
                            does_stockpile_exist = cursor.fetchone()
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
                    if does_stockpile_exist is not None:
                        query = f"UPDATE stockpiles_list SET `active` = %s WHERE `id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, ['0', does_stockpile_exist['id']])
                                connection.commit()
                                message += f"Stockpile **{x['name']}** has been deleted!\n"
                        except pymysql.err.ProgrammingError as except_detail:
                            message += "Error deleting stockpile in database, iv removed it from the discord but " \
                                       "cannot locate it in the database. This may or may not cause issues down the line. Error Code: CJKWt0Iy\n"
                            print(except_detail)
                    else:
                        message += "Error deleting stockpile in database, iv removed it from the discord but " \
                                   "cannot locate it in the database. This may or may not cause issues down the line. Error Code: K1QjAWIz\n"
                    stockpile_name = x['name']
                    del stock_data[counter]
                    await commands_channel.send(message)
                    await stockpile_data.edit(content=(code_warning + json.dumps(stock_data)))
                    await codes_channel.send(f"Stockpile {stockpile_name} DELETED!! by: {ctx.message.author.mention}")
                    await display_current_stockpiles(ctx, args, discord_data)

                counter += 1

            # Try to delete the add command the person posted, we don't really care if it fails.
            try:
                await ctx.message.delete()
            except:
                return

    else:
        await ctx.send(
            f"{ctx.message.author.mention} This discord is not setup with my yet. Use the command **!stock setup wizard** to start the process!\n"
            "After im setup you can add stockpiles for me to keep track of here!")
        return


async def create_stockpile_database(codes_channel, code_warning, discord_data, ctx):
    print("here")
    code_message = await codes_channel.send(code_warning)
    print(code_message)
    code_message_id = code_message.id

    query = f"UPDATE discords_list SET `stockpile_codes_message` = %s WHERE `id` = %s"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query, [code_message_id, discord_data['id']])
            connection.commit()
    except pymysql.err.ProgrammingError as except_detail:
        await ctx.send("Error setting Commands Channel, please try again. Error code: 38eF0eqo")
        print(except_detail)
        return ""
    return code_message_id


def check_ban_status(ctx):
    query = f"SELECT `status` FROM discords_list WHERE `discord_name`='{ctx.guild.name}'"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            discord_data = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return
    if discord_data is not None:
        if discord_data == 0:
            ctx.send(
                "Im sorry, your discord has been BANNED for use with stockpiler. If you think this is a mistake please reach out! support@stockpiler.net")
            return True
        else:
            return False


@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, CommandNotFound):
        return
    raise error


async def alert_command(ctx, args, discord_data=[]):
    command_channel = discord_data["commands_channel"]
    if len(args) == 2:
        await get_command_channel(ctx, command_channel).send(
            f"{ctx.message.author.mention} Alerts will ping users when a stockpile is running out of time.\n"
            "To configure an alert, use the command:\n"
            "**!stock setup alert add MinutesLeftInStockpile RoleToPing 'Stockpile Name'**\n"
            "Stockpile name is optional, if left out the alert will be set to all stockpiles related to this discord.\n"
            "Only set name if you want to set an alert for a particular stockpile only\n"
            "Exmaple command: **!stock setup alert add 600 everyone**\n"
            "Will ping @ everyone when the stockpile expires in 10 hours.\n"
            "**!stock setup alert add 120 leader 'Blemish'**\n"
            "Will ping anyone with the leader role when the Blemish and only the Blemish stockpile expires in 2 hours\n"
            "To see all active alerts for this discord use the command **!stock setup alert list**\n"
            "Delete an alert with **!stock setup alert del alertID**\n"
            "Alert IDs are shown with the command **!stock setup alert list**")
        return
    if len(args) == 3 and args[2].lower() == "list":
        query = f"SELECT * FROM discord_stockpiles WHERE `discords` = '{discord_data['id']}'"
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query)
                discord_stockpiles = cursor.fetchall()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return
        if len(discord_stockpiles) > 0:
            query = f"SELECT * FROM stock_alarms_config WHERE `discord` = '{discord_data['id']}'"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    discord_alarms = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            if len(discord_alarms) > 0:
                alert_list = ""
                for x in discord_alarms:
                    name = ""
                    if x['stockpile_id'] is not None:
                        for y in discord_stockpiles:
                            if x['stockpile_id'] == y['stockpile_id']:
                                name = y['stockpile_name']

                    time = x['minutes_left']
                    if time > 0:
                        time = f"{time} minutes ({round((time / 60), 2)} hours)"
                    role = x['roles_to_ping']

                    alert_list += f"- {name} Will alert when {time} is left and will ping the {role} Role. To delete alert use command: !stock setup alert del {x['id']}\n"
                await get_command_channel(ctx, command_channel).send(
                    f"{ctx.message.author.mention} The following alerts are configured for this discord:\n"
                    f"{alert_list}")
                return
            else:
                await get_command_channel(ctx, command_channel).send(
                    f"{ctx.message.author.mention} No alerts have been set yet. Use the command !stock setup alert for details on setting one.")
                return
        else:
            await get_command_channel(ctx, command_channel).send(
                f"{ctx.message.author.mention} I'm not monitoring any stockpiles for this discord. To add one, use the **!stock add code** given to you in the logs when you make a new stockpile on https://beta.stockpiler.net")
            return
    elif len(args) >= 3 and args[2].lower() == "add":
        if len(args) >= 5:
            if args[3].isnumeric():
                role_found = False
                role_request = args[4].lower()
                for x in ctx.guild.roles:
                    if role_request == "here":
                        role_found = True
                    elif role_request == "everyone":
                        role_found = True
                    elif role_request == "online":
                        role_found = True
                    elif x.name == args[4]:  # Want to maintain case sensitivity for other roles
                        role_found = True
                if not role_found:
                    await get_command_channel(ctx, command_channel).send(
                        f"{ctx.message.author.mention} I'm sorry, '{role_request}' is not a vaild role for this discord.\n"
                        f"Roles are Case-Sesitive. Admin and admin are not the same role.")
                    return
                else:
                    if len(
                            args) >= 6:  # There is a 6th argument, it should be the name of the stockpile for this alert. This one is optional, if left out the alert applys to all stockpiles for this discord.
                        if args[5][0] == "'" and args[5][-1] == "'":  # ensure the stockpile name is enclosed in '
                            stockpile_name_found = False
                            stockpile_id = ""
                            stockpile_name = ""
                            for x in discord_stockpiles:
                                if args[5].lower() == x['stockpile_name'].lower():
                                    stockpile_name_found = True
                                    stockpile_id = x['stockpile_id']
                                    stockpile_name = x['stockpile_name']
                            if stockpile_name_found:
                                query = f"INSERT INTO stock_alarms_config (discord, stockpile_id, minutes_left, roles_to_ping) VALUES (%s, %s, %s, %s)"
                                data = (discord_data['id'], stockpile_id, args[3], args[4])
                                try:
                                    connection.ping()
                                    with connection.cursor() as cursor:
                                        cursor.execute(query, data)
                                        connection.commit()
                                    await get_command_channel(ctx, command_channel).send(
                                        f"{ctx.message.author.mention} A ping has been configured for stockpile **{stockpile_name}** only. If the timers gets down to {args[3]} minutes or less ill ping the {args[4]} role.")
                                    return
                                except pymysql.err.ProgrammingError as except_detail:
                                    print(except_detail)
                                    return

                            else:  # the stockpile name they listed could not be found.
                                await get_command_channel(ctx, command_channel).send(
                                    f"{ctx.message.author.mention} You tried to set this alert for '**{args[5]}**' however this stockpile does not exist or is not connected to this discord. \n"
                                    f"Make sure you have the spelling exactly as you see in stockpiler, or make sure the stockpile is added to this discord.\n"
                                    f"The command for adding a stockpile to a discord is one of the first few logs shown when a stockpile is created on https://www.stockpile.net")
                                return

                        else:
                            await get_command_channel(ctx, command_channel).send(
                                f"{ctx.message.author.mention} You tried to set this alert for **{args[5]}** however you did not include the ''\n"
                                f"**!stock setup alert add {args[3]} {args[4]} {args[5]}**\n")
                            return
                    else:
                        query = f"INSERT INTO stock_alarms_config (discord, minutes_left, roles_to_ping) VALUES (%s, %s, %s)"
                        data = (discord_data['id'], args[3], args[4])
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, data)
                                connection.commit()
                            await get_command_channel(ctx, command_channel).send(
                                f"{ctx.message.author.mention} A ping has been configured for ALL Stockpiles connected to this discord. If any of their timers gets down to {args[3]} minutes or less ill ping the {args[4]} role.")
                            return
                        except pymysql.err.ProgrammingError as except_detail:
                            print(except_detail)
                            return


            else:
                await get_command_channel(ctx, command_channel).send(
                    f"{ctx.message.author.mention} Looks like you did NOT enter a number for minutes, please enter in minutes for how long till expire the alert should go off\n"
                    f"**!stock setup alert add MinutesLeftInStockpile RoleToPing**")
                return
        else:
            await get_command_channel(ctx, command_channel).send(
                f"{ctx.message.author.mention} Looks like your missing some paramaters for this command.\n"
                f"The command is **!stock setup alert add MinutesLeftInStockpile RoleToPing**\n"
                f"Use **!stock setup alert** to see examples. Error Code: E1Ob1SpL")
            return
    elif len(args) >= 3 and args[2].lower() == "del":

        if len(args) == 4:
            if args[3].isnumeric():
                query = f"SELECT * FROM stock_alarms_config WHERE `id` = %s"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query, [args[3]])
                        get_alerts = cursor.fetchone()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
                if get_alerts is not None:
                    if get_alerts["discord"] != discord_data["id"]:  # This should only happen when the person is
                        # trying to delete a alert that was made on another stockpile. Are they trying to hack?
                        await get_command_channel(ctx, command_channel).send(
                            f"{ctx.message.author.mention} An unexpected error has occured. Please contact stockpiler support and provide this error code support@stockpiler.net\n"
                            f"Error Code: pzQcmQbC")
                        # TODO add logging for hacking attempt
                        return
                    else:
                        query = "DELETE FROM stock_alarms_config WHERE id = %s"
                        data = (str(args[3]))
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, data)
                                connection.commit()
                                await get_command_channel(ctx, command_channel).send(
                                    f"{ctx.message.author.mention} Alert number {args[3]} has been removed.\n")
                                return
                        except pymysql.err.ProgrammingError as except_detail:
                            print(except_detail)
                            await get_command_channel(ctx, command_channel).send(
                                f"{ctx.message.author.mention} Unknown error when deleting that alert, " \
                                f"please try again. Error Code: TJNv007Y")
                            return
                else:  # they tried to delete a stockpile code that did not exist, maybe hacking?
                    await get_command_channel(ctx, command_channel).send(
                        f"{ctx.message.author.mention} An unexpected error has occured. Please contact stockpiler support and provide this error code support@stockpiler.net\n"
                        f"Error Code: J3SEuZ0m")
                    # TODO add logging for hacking attempt
                    return

            else:
                await get_command_channel(ctx, command_channel).send(
                    f"{ctx.message.author.mention} Thats not a correct delete code. You can see delete codes for alerts with !stock setup alert list\n"
                    f"**!stock setup alert del #** Error Code: YzeCr5lu")
                return
        else:
            await get_command_channel(ctx, command_channel).send(
                f"{ctx.message.author.mention} You did not enter a valid delete command. Its **!stock setup alert del #** Error Code: YzeCr5lu")
            return
    return


async def setup_command(ctx, args):
    banned = check_ban_status(ctx)
    if banned:
        return

    query = f"SELECT * FROM discords_list WHERE `discord_name` = '{ctx.guild.name}'"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            discord_data = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return




    log_channels = []
    if discord_data is None:
        query = f"INSERT INTO discords_list (discord_name, discord_id) VALUES (%s, %s)"
        data = [ctx.guild.name, ctx.guild.id]
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query, data)
                connection.commit()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return
        query = f"SELECT * FROM discords_list WHERE `discord_name` = '{ctx.guild.name}'"
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query)
                discord_data = cursor.fetchone()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return
        if len(args) <= 1 or args[1].lower() != "wizard":
            await ctx.send(
                f"{ctx.message.author.mention} Looks like this is your first time using Stockpiler for {ctx.guild.name}.\n"
                f"Use the command **!stock setup wizard**\n"
                f"To start the Setup Wizard to configure Stockpiler for use with this Discord.\n"
                f"Enter **!stock setup** again for a list of setup commands.")
            return

    if discord_data is not None:
        # ------------------------------------------------------------------
        if len(args) == 1 and args[0].lower() == "setup":
            message = "Stockpiler setup command list:\n" \
                      "**!stock setup** - Lists setup commands available.\n" \
                      "**!stock setup wizard** - Starts a first time walkthrough of setting up stockpiler.\n" \
                      "**!stock setup commands_channel #channel** - Sets what channel stockpiler will respond to when a " \
                      "!stock command is given. use the word **skip** in place of a #channel to unset this channel, " \
                      "as such stockpiler will reply to commands in whatever channel the command is given in.\n" \
                      "**!stock setup stock_codes_channel #channel** - Sets what channel stockpiler will respond to when a " \
                      "*!stock codes* command is given. This command lists all stockpile codes the requester has access to. " \
                      "Use the word **skip** in place of a #channel to unset this channel.\n" \
                      "**!stock setup log_channel #channel1 #channel2 #channel3** - You can set up to 10 channels with " \
                      "this command. Will set where stockpiler will post and stockpile changes that are set on the www.stockpiler.net website." \
                      "Use the word **skip** in place of a #channel1 to unset ALL channel.\n" \
                      "**!stock setup admin_role** - The role that can use the !stock setup command as well as a few others. You can mention the role or just type it out, if you type it make sure its 100% correct.\n" \
                      "**!stock setup access_role** - The role that a person must have to view the guild stockpiles on the website.\n" \
                      "**!stock setup status** - See what all settings are set to currently.\n" \
                      "**!stock setup alert** - Configure alerts for when a stockpiles timer gets low.\n"

            await get_command_channel(ctx).send(message)
            return

        if len(args) == 2 and args[1].lower() == "wizard":
            query = f"UPDATE discords_list SET `setup_status` = %s WHERE `id` = %s"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query, [1, discord_data['id']])
                    connection.commit()
            except pymysql.err.ProgrammingError as except_detail:
                await ctx.send("Error setting Commands Channel, please try again. Error code: 38eF0eqo")
                print(except_detail)
                return
            discord_data['setup_status'] = 1
            await setup_process(0, args, ctx, discord_data) # 0 is saying to run through all settings checking each one is set.


        if len(args) >= 2 and args[1] == "commands_channel":
            await setup_process(1, args, ctx, discord_data) # 1 stands for set command_channel
            return

        if len(args) >= 2 and args[1].lower() == "stock_codes_channel":
            await setup_process(2, args, ctx, discord_data)  # 2 stands for set what channel is used to track and list stockpile codes
            return

        if len(args) >= 2 and args[1] == "log_channel":
            await setup_process(3, args, ctx, discord_data)  # 3 stands for set what channel is used to post stockpile change logs to.
            return

        if len(args) >= 2 and args[1] == "admin_role":
            await setup_process(4, args, ctx, discord_data)
            return

        if len(args) >= 2 and args[1] == "access_role":
            await setup_process(5, args, ctx, discord_data)
            return


        if len(args) >= 2 and args[1] == "status":
            await setup_process(6, args, ctx, discord_data)
            return

        if len(args) >= 2 and args[1].lower() == "alert":
            #TODO review
            await alert_command(ctx, args, discord_data)
            return
    else:
        await ctx.send(
            f"{ctx.message.author.mention} Unknown error with setup command, please try again. Error Code: hEwbE9V0")
        return



async def setup_process(progress, args, ctx, discord_data):
    total_wizard_steps = 5

    if discord_data['setup_status'] == 1:
        progress = 0
    # --- Commands Channel setup --- #
    if progress == 1 or progress == 0:
        output_message = output_message = ctx.message.author.mention + " "
        if progress == 0:
            output_message += f"*Stockpiler Setup Wizard - Step 1/{total_wizard_steps}*\n\n"
        command_channel = discord_data["commands_channel"]
        if command_channel is None or len(command_channel) == 0 or args[1] == "commands_channel":
            if len(args) >= 3 and args[2] == "skip":
                query = f"UPDATE discords_list SET `commands_channel` = %s WHERE `id` = %s"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query, ['skip', discord_data['id']])
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    await ctx.send("Error setting Commands Channel, please try again. Error code: joYUlixl")
                    print(except_detail)
                    return
                output_message +="Skipping Command Channel. Ill reply wherever you post.\n"
                "You can set this channel later with the command.\n"
                "**!stock setup commands_channel #channel**"
                await ctx.send(output_message)

            else:
                if len(args) >= 3 and args[1] == "commands_channel":
                    # Can't find a was to remove the discord channel tags other then some replaces...
                    channel_id = str(args[2]).replace("<#", "").replace(">", "")
                    # Verifying the chat channel exists.
                    if channel_id.strip().isdigit() and len(str(int(channel_id))) > 10:
                        channel = bot.get_channel(int(channel_id))
                    else:
                        channel = discord.utils.get(bot.get_all_channels(), guild__name=ctx.guild.name,
                                                    name=channel_id)
                    if channel is None:
                        output_message += f"'{args[2]}' is not a valid channel. Did you misstype? Error Code: 1q7y6otf"
                        await ctx.send(output_message)
                        return
                    else:
                        #updating the database with the new channel
                        query = f"UPDATE discords_list SET `commands_channel` = %s WHERE `id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, [channel.id, discord_data['id']])
                                connection.commit()
                        except pymysql.err.ProgrammingError as except_detail:
                            await ctx.send("Error setting Commands Channel, please try again. Error code: 38eF0eqo")
                            print(except_detail)
                            return
                        if command_channel is None or len(command_channel) == 0:
                            output_message += f"Commands channel set to: **#{channel.name}**"
                            await ctx.send(output_message)

                        else:
                            output_message += f"Commands channel changed FROM: **#{bot.get_channel(int(command_channel)).name}** TO: **#{channel.name}**"
                            await ctx.send(output_message)


                else:
                    output_message += "What text channel would you like me to reply to commands in?\n"\
                                      "Reply with **!stock setup commands_channel #channel**\n\n"\
                                      "If you set a channel, you can type commands in any channel you like however i will only respond in the channel specified.\n"\
                                      "You can skip this part of the setup if you like with the command **!stock setup commchannel skip**\n"\
                                      "Skipping Commands Channel setup means i will reply in whatever channel someone uses.\n"
                    await ctx.send(output_message)
                    return
    # --- Stockpile Codes Channel setup --- #
    if progress == 2 or progress == 0:
        output_message = ctx.message.author.mention + " "
        if progress == 0:
            output_message += f"*Stockpiler Setup Wizard - Step 2/{total_wizard_steps}*\n\n"
        codes_channel = discord_data["stockpile_code_channel"]
        if codes_channel is None or len(codes_channel) == 0 or args[1] == "stock_codes_channel":
            if len(args) >= 3 and args[2] == "skip":
                query = f"UPDATE discords_list SET `stockpile_code_channel` = %s WHERE `id` = %s"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query, ['skip', discord_data['id']])
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    await get_command_channel(ctx).send("Error setting Stockpile Codes Channel, please try again. Error code: EbEDnDBd")
                    print(except_detail)
                    return
                output_message += "Skipping Stockpile Codes Channel. Ill reply with stockpile codes wherever someone posts post.\n"
                "You can set this channel later with the command.\n"
                "**!stock setup stock_codes_channel #channel**"
                await get_command_channel(ctx).send(output_message)

            else:
                if len(args) >= 3 and args[1] == "stock_codes_channel":
                    # Can't find a was to remove the discord channel tags other then some replaces...
                    channel_id = str(args[2]).replace("<#", "").replace(">", "")
                    # Verifying the chat channel exists.
                    if channel_id.strip().isdigit() and len(str(int(channel_id))) > 10:
                        channel = bot.get_channel(int(channel_id))
                    else:
                        channel = discord.utils.get(bot.get_all_channels(), guild__name=ctx.guild.name,
                                                    name=channel_id)
                    if channel is None:
                        output_message += f"'{args[2]}' is not a valid channel. Did you misstype? Error Code: PTt5MJPH"
                        await get_command_channel(ctx).send(output_message)
                        return
                    else:
                        # updating the database with the new channel
                        query = f"UPDATE discords_list SET `stockpile_code_channel` = %s WHERE `id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, [channel.id, discord_data['id']])
                                connection.commit()
                        except pymysql.err.ProgrammingError as except_detail:
                            await get_command_channel(ctx).send("Error setting Stockpile Codes Channel, please try again. Error code: iQ0xCG8a")
                            print(except_detail)
                            return
                        if codes_channel is None or len(codes_channel) == 0:
                            output_message += f"Stockpile Codes channel set to: **#{channel.name}**"
                            await get_command_channel(ctx).send(output_message)

                        else:
                            output_message += f"Stockpile Codes Channel changed FROM: **#{bot.get_channel(int(codes_channel)).name}** TO: **#{channel.name}**"
                            await get_command_channel(ctx).send(output_message)


                else:
                    output_message += "What text channel would you like me to post stockpile codes in?\n" \
                                      "Reply with **!stock setup stock_codes_channel #channel**\n\n" \
                                      "If you set a channel, you can type **!stock codes** in any channel you like " \
                                      "however i will only respond with the stockpile codes you have access to in the specified channel.\n" \
                                      "You can skip this part of the setup if you like with the command **!stock setup stock_codes_channel skip**\n" \
                                      "Skipping Commands Channel setup means i will reply in whatever channel someone uses.\n" \
                                      "\nSkipping the Commands Channel setup is NOT recomended!"
                    await get_command_channel(ctx).send(output_message)
                    return
    # --- Log Channel setup --- #
    if progress == 3 or progress == 0:
        output_message = output_message = ctx.message.author.mention + " "
        if progress == 0:
            output_message += f"*Stockpiler Setup Wizard - Step 3/{total_wizard_steps}*\n\n"
        log_channel = discord_data["channel_name"]
        if log_channel is None or len(log_channel) == 0 or args[1] == "log_channel":
            if len(args) >= 3 and args[2] == "skip":
                query = f"UPDATE discords_list SET `channel_name` = %s WHERE `id` = %s"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query, ['skip', discord_data['id']])
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    await get_command_channel(ctx).send("Error setting Log Channel, please try again. Error code: jFmLBAZG")
                    print(except_detail)
                    return
                output_message += "Skipping Log Channel. I will not update the discord when stockpile changes occur.\n"
                "You can set this channel later with the command.\n"
                "**!stock setup log_channel #channel1 #channel2** ect.. up to 10 channels"
                await get_command_channel(ctx).send(output_message)

            else:
                if len(args) >= 3 and args[1] == "log_channel":
                    # Can't find a was to remove the discord channel tags other then some replaces...
                    channel_id_list = []
                    for x in args:
                        if "<#" in str(x):
                            ids = str(x).split("<#")
                            if len(ids) > 0:
                                for g in ids:
                                    temp = g.replace("<#", "").replace(">", "")
                                    if len(temp) > 0:
                                        channel_id_list.append(temp)
                            else:
                                temp = x.replace("<#", "").replace(">", "")
                                if len(temp) > 0:
                                    channel_id_list.append(temp)
                        else:
                            names = str(x).split(" ")
                            for k in names:
                                if k == 'setup' or 'log_channel ':
                                    continue
                                channel_id_list.append(k.replace("#", ""))
                    if len(channel_id_list) > 10:
                        await get_command_channel(ctx).send("")
                    error_message = ""
                    channel_names = ""
                    for channelid in channel_id_list:
                        # Verifying the chat channel exists.
                        if channelid.strip().isdigit() and len(str(int(channelid))) > 10:
                            channel = bot.get_channel(int(channelid))
                            if channel is None:
                                error_message += "One of the channels in this list is invald. Error code: zAAtLUW8\n"
                            else:
                                channel_names += f"{channel.mention} "
                        else:
                            channel = discord.utils.get(bot.get_all_channels(), guild__name=ctx.guild.name,
                                                        name=channelid)
                            if channel is None:
                                error_message += f"{channelid} is not a valid channel, try linking the channel with a #. Error code: 5hiPCKXj\n"
                            else:
                                channel_names += f"{channel.mention} "

                    if len(error_message) > 0:
                        output_message += error_message
                        await get_command_channel(ctx).send(output_message)
                        return
                    else:
                        if len(channel_id_list) == 0:
                            output_message += "Error saving list of channels, please try again. Error Code: 9t5sXL7i"
                            await get_command_channel(ctx).send(output_message)
                            return
                        update_string = ""
                        for y in channel_id_list:
                            update_string += f"{y},"
                        update_string = update_string[:-1]
                        # updating the database with the new channel
                        query = f"UPDATE discords_list SET `channel_name` = %s WHERE `id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, [update_string, discord_data['id']])
                                connection.commit()
                        except pymysql.err.ProgrammingError as except_detail:
                            await get_command_channel(ctx).send("Error setting Log Channel, please try again. Error code: pPWDABNc")
                            print(except_detail)
                            return

                        output_message += f"Stockpile Log channels set to the following: {channel_names}"
                        await get_command_channel(ctx).send(output_message)
                else:
                    output_message += "What text channel would you like me to Stockpile Changes to?\n" \
                                      "Reply with **!stock setup log_channel #channel1 #channel2** ect.. up to 10 channels\n\n" \
                                      "You can list 1 or up to 10 channels, when someone makes changes on www.stockpiler.net the changes will be posted to the selected channels.\n" \
                                      "You can skip this part of the setup if you like with the command **!stock setup log_channel skip**\n" \
                                      "Skipping Log Channel setup means i will not update the discord when stock changes are made."
                    await get_command_channel(ctx).send(output_message)
                    return
    # --- Admin Role setup --- #
    if progress == 4 or progress == 0:
        output_message = ctx.message.author.mention + " "
        if progress == 0:
            output_message += f"*Stockpiler Setup Wizard - Step 4/{total_wizard_steps}*\n\n"
        admin_role = discord_data["admin_stockpiles_role"]
        if admin_role is None or len(admin_role) == 0 or args[1] == "admin_role":
            if len(args) >= 3 and args[1] == "admin_role":
                guild_roles = ctx.guild.roles
                role = None
                # Can't find a was to remove the discord channel tags other then some replaces...
                role_id = str(args[2]).replace("<@&", "").replace(">", "")
                # Verifying the chat channel exists.

                if role_id.strip().isdigit() and len(str(int(role_id))) > 10:
                    for g in guild_roles:
                        if int(role_id) == g.id:
                            role = g
                else:
                    for g in guild_roles:
                        if role_id == g.name:
                            role = g

                if role is None:
                    output_message += f"'{args[2]}' is not a valid role. Did you misstype? Roles are case-sensitive Error Code: npq3MxzC"
                    await get_command_channel(ctx).send(output_message)
                    return
                else:
                    # updating the database with the new channel
                    query = f"UPDATE discords_list SET `admin_stockpiles_role` = %s WHERE `id` = %s"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query, [role.id, discord_data['id']])
                            connection.commit()
                    except pymysql.err.ProgrammingError as except_detail:
                        await get_command_channel(ctx).send("Error setting Admin Role, please try again. Error code: wrbCEIcp")
                        print(except_detail)
                        return
                    output_message += f"Admin role changed to: {role.name}"
                    await get_command_channel(ctx).send(output_message)



            else:
                output_message += "What role can use the !stock setup command?\n" \
                                  "Reply with **!stock setup admin_role @role**\n\n" \
                                  "This also effects who can setup low stockpile alarms (!stock setup alert)\n" \
                                  "@ ing the role is best, but you can also just type the roles name, it is case sesitive.\n" \
                                  "People with the Administorators role are not effected by this setting."
                await get_command_channel(ctx).send(output_message)
                return
    # --- Access Role --- #
    if progress == 5 or progress == 0:
        output_message = ctx.message.author.mention + " "
        if progress == 0:
            output_message += f"*Stockpiler Setup Wizard - Step 5/{total_wizard_steps}*\n\n"
        access_role = discord_data["view_stockpiles_role"]
        if access_role is None or len(access_role) == 0 or args[1] == "access_role":
            if len(args) >= 3 and args[1] == "access_role":
                guild_roles = ctx.guild.roles
                role = None
                # Can't find a was to remove the discord channel tags other then some replaces...
                role_id = str(args[2]).replace("<@&", "").replace(">", "")
                # Verifying the chat channel exists.

                if role_id.strip().isdigit() and len(str(int(role_id))) > 10:
                    for g in guild_roles:
                        if int(role_id) == g.id:
                            role = g
                else:
                    for g in guild_roles:
                        if role_id == g.name:
                            role = g

                if role is None:
                    output_message += f"'{args[2]}' is not a valid role. Did you misstype? Roles are case-sensitive Error Code: jpQApjhr"
                    await get_command_channel(ctx).send(output_message)
                    return
                else:
                    # updating the database with the new channel
                    query = f"UPDATE discords_list SET `view_stockpiles_role` = %s WHERE `id` = %s"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query, [role.id, discord_data['id']])
                            connection.commit()
                    except pymysql.err.ProgrammingError as except_detail:
                        await get_command_channel(ctx).send("Error setting Admin Role, please try again. Error code: 0yAQU5A1")
                        print(except_detail)
                        return
                    output_message += f"Access role changed to: {role.name}"
                    await get_command_channel(ctx).send(output_message)
                    discord_data['view_stockpiles_role'] = role.id



            else:
                output_message += "What role can view stockpiles on the website?\n" \
                                  "Reply with **!stock setup access_role @role**\n\n" \
                                  "@ ing the role is best, but you can also just type the roles name, it is case sesitive.\n" \
                                  "People with the Administorators role are not effected by this setting."
                await get_command_channel(ctx).send(output_message)
                return
    # --- Check Setup Status --- #
    if progress == 6 or progress == 0:
        output_message = ctx.message.author.mention + " \n"
        if progress == 0:
            complete_message = f"Stockpiler Setup **COMPLETE!**\n\n"
            await get_command_channel(ctx).send(complete_message)
            query = f"UPDATE discords_list SET `setup_status` = %s WHERE `id` = %s"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query, [5, discord_data['id']])
                    connection.commit()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
        commands_channel = discord_data['commands_channel']
        codes_channel = discord_data['stockpile_code_channel']
        log_channels = discord_data['channel_name']
        admin_role = discord_data['admin_stockpiles_role']
        access_role = discord_data['view_stockpiles_role']
        output_message += "**Stockpiler Setup Status**: \n\n"
        # --- Commands channel --- #
        output_message += "Commands Channel: "
        if commands_channel is None or len(commands_channel) == 0:
            output_message += "~~Not Set~~\n"
        else:
            channel = bot.get_channel(int(commands_channel))
            if channel is not None:
                output_message += f"{channel.mention}\n"
            else:
                output_message += "ERROR Verifying channel.\n"
        channel = None

        # --- Codes channel --- #
        output_message += "Stockpile Codes Channel: "
        if codes_channel is None or len(codes_channel) == 0:
            output_message += "~~Not Set~~\n"
        else:
            channel = bot.get_channel(int(codes_channel))
            if channel is not None:
                output_message += f"{channel.mention}\n"
            else:
                output_message += "ERROR Verifying channel.\n"
        channel = None

        # --- Logging channels --- #
        output_message += "Stockpile Changes Log Channels: \n"
        if log_channels is None or len(log_channels) == 0:
            output_message += "~~Not Set~~\n"
        else:
            temp = log_channels.split(",")
            for d in temp:
                channel = bot.get_channel(int(d))
                if channel is not None:
                    output_message += f"{channel.mention}\n"
                else:
                    output_message += "ERROR Verifying channel.\n"
        channel = None

        # --- Admin Role --- #
        output_message += "Stockpiler Admin Role: "
        if admin_role is None or len(admin_role) == 0:
            output_message += "~~Not Set~~\n"
        else:
            role = None
            for g in ctx.guild.roles:
                if int(admin_role) == g.id:
                    role = g
            if role is not None:
                output_message += f"**{role.name}**\n"
            else:
                output_message += "ERROR Verifying role.\n"
        channel = None

        # --- Access Role --- #
        output_message += "Stockpiler Access Role: "
        if access_role is None or len(access_role) == 0:
            output_message += "~~Not Set~~\n"
        else:
            role = None
            for g in ctx.guild.roles:
                if int(admin_role) == g.id:
                    role = g
            if role is not None:
                output_message += f"**{role.name}**\n"
            else:
                output_message += "ERROR Verifying role.\n"

        output_message += f"\n Use command **!stock setup** to see a list of commands to change these settings."
        await get_command_channel(ctx).send(output_message)
        return


@bot.command(name='help', help='Provides information on the bot.')
async def helpc(ctx):
    await get_command_channel(ctx).send(f"The help command for Stockpiler is !stock help")


@bot.command(name='stock',
             help='Provides information on current stock levels for the given stockpile. use !stock or !stock Stockpile Name')
async def stock_command(ctx, *args):
    if len(args) == 0:
        await get_command_channel(ctx).send(f"Command **!Stock** ran.\n\n"
                                            f"**--This command has changed!--**\n\n"
                                            f"If you wish to see all stockpiles related to this discord use the command !stock list\n"
                                            f"You can filter whats shown with a filter word as well\n"
                                            f"Filter words are: Light, Heavy, Logi, Tools, Med, Vehicals, Large, Shirts\n"
                                            f"Example, to only show vehicals and heavy ammo in stock use !stock list vehicals heavy\n"
                                            f"We also have a special command that only shows these 2. !stock armor\n"
                                            f"You can still use '!stock Stockpile Name' to show the stock for just that stockpile.\n"
                                            f"**BETA**: https://beta.stockpiler.net\n"
                                            f"If BETA has issues you can fall back to the main site.\n"
                                            f"**Main Website**: https://www.stockpiler.net\n"
                                            f"\n"
                                            f"To create a stockpile hotlink, add the stockpile ID to the end of this URL:\n"
                                            f"**Current command list:**\n"
                                            f"!stock\n"
                                            f"!stock list\n"
                                            f"!stock armor\n"
                                            f"!stock setup\n"
                                            f"!stock add\n"
                                            f"!stock del\n"
                                            f"!stock rename\n")
        '''
        try:
            await ctx.message.delete()
        except discord.NotFound as response:
            print(f"Unable to delete message, not found: {response}")

        '''
        return
    async with ctx.channel.typing():
        stockpile = ""
        authorized = []
        args_list = ""
        list_filtered = []
        ran = False
        print(f"Stock command ran for guild '{ctx.guild.name}' : !stock {args}")
        discord_stockpiles = []

        if len(args) > 0:
            args_string = str(args[0]).lower()
            if args_string == "list" or args_string == "armor":
                query = f"SELECT * FROM discord_stockpiles"
                failed_filter_check = False
                if args_string == "list":
                    for x in args:
                        if not failed_filter_check:
                            if x == "list":
                                pass
                            else:
                                if x == "light" or x == "heavy" or x == "logi" or x == "tools" \
                                        or x == "med" or x == "vehicals" or x == "large" \
                                        or x == "shirts":
                                    list_filtered.append(x)
                                else:
                                    failed_filter_check = True
                    if failed_filter_check:
                        await get_command_channel(ctx).send(f"Command **!Stock list** ran.\n"
                                                            f"ERROR: Looks like you entered something that was not a filter\n"
                                                            f"If you wish to see all stockpiles related to this discord use the command !stock list\n"
                                                            f"You can filter whats shown with a filter word as well\n"
                                                            f"Filter words are: Light, Heavy, Logi, Tools, Med, Vehicals, Large, Shirts\n"
                                                            f"Example, to only show vehicals and heavy ammo in stock use !stock list vehicals heavy\n"
                                                            f"We also have a special command that only shows these 2. !stock armor\n")
                        return
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query)
                            discord_stockpiles = cursor.fetchall()
                        ran = True
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
                        '''
                        try:
                            await ctx.message.delete()
                        except discord.NotFound as response:
                            print(f"Unable to delete message, not found: {response}")
                        return
                        '''
                elif args_string == "armor":
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query)
                            discord_stockpiles = cursor.fetchall()
                        ran = True
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
            elif args_string == "setup":
                await setup_command(ctx, args)
                return
            elif args_string == "add":
                await add_stockpile(ctx, args)
                return
            elif args_string == "del":
                await del_stockpile(ctx, args)
                return
            elif args_string == "rename":
                await rename_stockpile(ctx, args)
                return
            elif args_string == "codes":
                await display_current_stockpiles(ctx, args)
                return
            elif args_string == "serverinfo":
                await server_info(ctx)
                return
            elif args_string == "list":
                stockpile_data = await selected_channel.fetch_message(discord_data['stockpile_codes_message'])
                stock_data = json.loads(message)
            else:
                for x in args:
                    args_list += x + ' '

                stockpile = args_list
                query = "SELECT * FROM discord_stockpiles WHERE stockpile_name = '" + stockpile + "'"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query)
                        discord_stockpiles = cursor.fetchall()
                        ran = True
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return
        if len(discord_stockpiles) > 0:
            discord_ids = []
            discord_data = []
            query = f"SELECT * FROM discords_list WHERE `discord_name` = '{ctx.guild.name}'"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    discord_list_data = cursor.fetchone()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return

            query = f"SELECT * FROM stockpiles_list"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    stockpile_list = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            for x in discord_stockpiles:
                passed = False
                for k in stockpile_list:
                    if k['last_addition'] is not None:
                        if k['internal_name'] == x['stockpile_id']:
                            EST = pytz.timezone(mytimezone)
                            now = datetime.now(EST)
                            server_timez = pytz.timezone(server_time)
                            timeleft = now - server_timez.localize(k['last_addition'])
                            minutes_left = math.floor(timeleft.total_seconds() / 60)
                            if minutes_left < 2900:
                                passed = True
                if not passed:
                    continue
                else:
                    if int(discord_list_data['id']) == int(x['discords']):
                        authorized.append([x['stockpile_id'], x['stockpile_name']])
        else:
            await get_command_channel(ctx).send(
                f"Unable to locate a stockpile by the name {stockpile} that this discord has access to. Perhaps "
                f"you have the wrong stockpile name or that stockpile is not added to this discord. Error Code: PefyJYlM")
            return
        string = f"Command **!Stock** ran. !help stock for details about this command.\n"
        if len(authorized) > 0:
            temp = ""
            name_single = ""
            if len(authorized) == 1:
                name_single = authorized[0][1]

            for lists in authorized:  # Go through each stockpile we are authed for and get and post stock
                temp += f"`stockpile_id` = '{lists[0]}' OR "

            temp = temp[:-4]
            string_additions = ""
            if len(args) > 0 and str(args[0]).lower() == "armor":
                string_additions += " AND ("
                for x in item_list.values():
                    if x["type"] == "vehicals" or x["type"] == "heavy" or x["type"] == "vehical_crates":
                        string_additions += f"`item` = '{x['pcname']}' OR "
                string_additions = string_additions[:-4]
                string_additions += ")"
            elif len(list_filtered) > 0:
                for z in list_filtered:
                    string_additions += " AND ("
                    for x in item_list.values():
                        if z == "vehicals":
                            if x["type"] == z or x["type"] == "vehical_crates":
                                string_additions += f"`item` = '{x['pcname']}' OR "
                        else:
                            if x["type"] == z:
                                string_additions += f"`item` = '{x['pcname']}' OR "
                    string_additions = string_additions[:-4]
                    string_additions += ")"

            query = f"SELECT * FROM stock_stock WHERE ({temp}) AND count > 0{string_additions}"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    stock = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            if len(stock) > 0:

                discord_name = discord_list_data["discord_name"]
                randnum = randint(0, 99999)
                name = ""
                if name_single == "":
                    name = discord_name + " Stockpiles"
                else:
                    name = name_single
                stockimg = generate_stock_image({"stockpile_id": discord_name, "id": randnum}, stock, name, True)
                await get_command_channel(ctx).send(file=discord.File(stockimg))
                '''
                await ctx.message.delete()
                '''
            else:
                await get_command_channel(ctx).send(
                    f"Unable to locate any stock matching those items thats associated with this discord. Error Code: 9txTBM5V")
                return


        else:
            if len(args) > 0:
                message = f"Unable to locate a stockpile by the name {stockpile} that this discord has access to. " \
                          f"Perhaps you have the wrong stockpile name or that stockpile is not added to this discord. " \
                          f"Error Code: KGgCWOQZ"
            else:
                message = "Iv got no stockpiles tied to this discord this war. You can make one on https://www.stockpiler.net"
            await get_command_channel(ctx).send(message)
            '''
            await ctx.message.delete()
            '''
            return


async def server_info(ctx):
    await ctx.send(f"**Server Name: **{ctx.guild.name}\n"
             f"**Server ID: **{ctx.guild.id}\n")

def generate_stock_image(log, stock, stockpile_name="", stock_command=False):
    stock_array = {}
    largest_total = 0
    stock_command_array = []
    for x in stock:
        if x['item'] in stock_array:
            stock_array[x["item"]] += x["count"]
        else:
            stock_array[x["item"]] = x["count"]
        stock_command_array.append([x["item"], x["count"], x['stockpile_id']])

        if x["count"] > largest_total:
            largest_total = x["count"]  # For use in some width stuff later
    additions = []
    removals = []

    correction = False
    if not stock_command:
        if log["item"] is None:
            return
        if log["log_type"] == "correction":
            correction = True
        for x in log["item"].split("|"):
            y = x.split("=")
            if int(y[1]) > 0:
                additions.append(y)
            else:
                removals.append(y)
    else:
        additions = stock_command_array

    if stock_command:
        query = f"SELECT * FROM stockpiles_list"
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query)
                stock_list = cursor.fetchall()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return
        if stock_list is not None and len(stock_list) > 0:
            temp = {}
            for y in stock_list:
                if y['internal_name'] not in temp:
                    temp[y['internal_name']] = y
            stock_list = temp

    imgwidth = 440
    lengthmax = 400  # Width will be a static 400px but legnth may change based on how much information there is.
    current_length = 0
    items_per_row = 3
    img = Image.new('RGBA', (imgwidth, lengthmax), color=(0, 0, 0))

    d = ImageDraw.Draw(img)
    current_length += 25  # for title bar
    d.rectangle([(0, 0), (imgwidth, current_length)], outline=(80, 80, 80), fill=(80, 80, 80))

    current_length += 5  # little padding

    fnt = ImageFont.truetype('Oswald-Regular.ttf', 15)
    fnt_label = ImageFont.truetype('Oswald-Regular.ttf', 13)
    d.text((15, 2), stockpile_name, fill=(255, 255, 255), font=fnt)
    if not stock_command:
        user = log["notes"].split("<b>")[1].split("</b>")[0]
        d.text((int(imgwidth / 2), 2), user, fill=(231, 124, 72), font=fnt)  # Username of the person in RED

    # running a caculation to find what our count box width should be.
    digit_count = 0
    while (largest_total > 0):
        largest_total = largest_total // 7
        digit_count += 1
    count_box_width = digit_count * 25  # for every digit add 20px to the width

    looper = 1
    run_Array = []
    while looper < 3:
        ran = False
        img = check_length(current_length, imgwidth, img, 90)
        d = ImageDraw.Draw(img)

        add = False
        if looper == 1:
            if len(additions) > 0:
                # Additions
                if not correction:
                    add_icon = Image.open('./imgs/icons/add_icon.png')
                    add_icon = add_icon.resize((30, 30))
                run_Array = additions
                add = True
        else:
            if len(removals) > 0:
                # Subtractions
                if not correction:
                    add_icon = Image.open('./imgs/icons/remove_icon.png')
                    add_icon = add_icon.resize((30, 30))
                run_Array = removals
            else:
                looper += 1
                continue
        if correction:
            add_icon = Image.open('./imgs/icons/correction_black.png')
            add_icon = add_icon.resize((73, 30))
        if len(run_Array) > 0:
            ran = True
            if not stock_command:
                img.paste(add_icon, [int(imgwidth / 2), current_length])  # In the center
                current_length += 35  # plus image + padding
            else:
                current_length += 5
            counter = 0
            fnt = ImageFont.truetype('Oswald-Regular.ttf', 20)

            if stock_command:
                list_font = ImageFont.truetype('Oswald-Regular.ttf', 10)
                stock_data = {}
                for o in run_Array:
                    if o[0] in stock_data:
                        stock_data[o[0]]['count'] += o[1]
                        stock_data[o[0]]['list'].append([o[2], o[1]])
                    else:
                        stock_data[o[0]] = {}
                        stock_data[o[0]]['count'] = o[1]
                        stock_data[o[0]]['list'] = []
                        stock_data[o[0]]['list'].append([o[2], o[1]])

            ran_already = []
            for x in run_Array:
                if stock_command:
                    if len(x) == 0 or len(x[0] == 0):
                        print(f"Error in logs datastream. Aborting.. {x}...{run_Array}")
                        continue
                    if x[0] in ran_already:
                        continue
                    else:
                        ran_already.append(x[0])
                img_loc = ""
                noimg = False
                try:
                    f = open(f"./imgs/icons/items/{x[0]}.png")
                    img_loc = f"./imgs/icons/items/{x[0]}.png"
                except IOError:
                    img_loc = f"./imgs/icons/items/na.png"
                    noimg = True

                item_icon = Image.open(img_loc)
                item_icon = item_icon.resize((50, 50))
                if counter == items_per_row:  # only 3 items per row
                    counter = 0
                    current_length += 70  # height per row
                    img = check_length(current_length, imgwidth, img, 80)
                    d = ImageDraw.Draw(img)

                current_width = 10
                if counter > 0:
                    current_width = (10 + (counter * int(imgwidth / items_per_row)))
                #  trying to change transparent pixels to black
                newImage = []
                for item in item_icon.getdata():
                    if item[3] == 255:
                        newImage.append((item[0], item[1], item[2]))
                    else:
                        newImage.append(item)

                item_icon.putdata(newImage)

                #  end

                img.paste(item_icon, [current_width, current_length])  # In the center

                if noimg:
                    d.text((current_width + 5, (current_length - 20)), item_list[x[0]]["name"], fill=(255, 255, 255),
                           font=fnt_label)  # counts / total
                bounding_box = [((current_width + 55), (current_length + 2)),
                                ((current_width + (55 + count_box_width)), (current_length + 40))]

                d.rectangle(bounding_box, fill=(80, 80, 80))  # Gray background for counts
                string_extra = ""
                if add and not stock_command:
                    string_extra = "+"
                string = ""
                if stock_command:
                    string = f"{stock_array[x[0]]}"
                else:
                    string = f"{string_extra}{x[1]}/{stock_array[x[0]]}"
                d.text(((bounding_box[0][0] + 2), (bounding_box[0][1] + 5)), string, fill=(255, 255, 255),
                       font=fnt)  # counts / total
                counter += 1
                if stock_command and len(stock_list) > 0:
                    temp_length = bounding_box[0][1] - 2
                    name_limit = 10
                    startingpoint = 0
                    if int(stock_array[x[0]]) > 9 and int(stock_array[x[0]]) < 99:
                        name_limit = 8
                        startingpoint += 6
                    elif int(stock_array[x[0]]) > 99 and int(stock_array[x[0]]) < 999:
                        name_limit = 6
                        startingpoint += 12
                    elif int(stock_array[x[0]]) > 999 and int(stock_array[x[0]]) < 9999:
                        name_limit = 3
                        startingpoint += 18
                    elif int(stock_array[x[0]]) > 9999:
                        continue
                    limiter = 0
                    sorted_stock_list = sorted(stock_data[x[0]]['list'], key=lambda stock: stock[1])
                    sorted_stock_list.reverse()
                    for g in sorted_stock_list:
                        limiter += 1
                        if limiter > 3:
                            continue
                        name = ''
                        if g[0] in stock_list:
                            name = stock_list[g[0]]['name']
                            if len(name) > name_limit:
                                name = name[:name_limit]
                        d.text(((bounding_box[0][0] + 20 + startingpoint), (temp_length)), f"{name}: {g[1]}",
                               fill=(255, 255, 255), font=list_font)  # counts / total
                        temp_length += 9
                    if limiter > 3:
                        d.text(((bounding_box[0][0] + 20 + startingpoint), (temp_length)), f"+{limiter} more..",
                               fill=(255, 255, 255), font=list_font)  # counts / total

            #
        looper += 1
        if ran:
            current_length += 80  # next row plus a little padding between addition and subtration

    # current_length -= 60  # since the last act was to add 80px lets cut that down a little

    fnt = ImageFont.truetype('Oswald-Regular.ttf', 10)
    d.text((150, (current_length - 15)), "beta.stockpiler.net", fill=(255, 255, 255),
           font=fnt)
    current_length += 15
    d.rectangle([(0, 0), (imgwidth, current_length)], outline=(80, 80, 80), width=2)  # Slight gray boarder
    img = img.crop((0, 0, imgwidth, current_length))
    savestring = f'.\stockpiler_bot_data\images\stock_changes\logimg_{log["stockpile_id"]}_{log["id"]}.png'
    img.save(savestring)
    # img.show()
    return savestring


def check_length(current_length, imgwidth, img, add_to_end_px):
    if (current_length + add_to_end_px) > img.height:
        newimg = Image.new('RGBA', (imgwidth, (current_length + add_to_end_px)), color=(0, 0, 0))
        loc = [int(img.width / 2), int(img.height / 2)]
        loc = [0, 0]
        newimg.paste(img, loc)
        return newimg
    else:
        return img


def get_command_channel(ctx, string=""):
    # string can be the name of a channel, will return None if not found
    if string == "":
        query = "SELECT commands_channel FROM discords_list WHERE discord_name = '" + str(ctx.guild.name) + "'"
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query)
                channel_name = cursor.fetchone()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return ctx
        if channel_name['commands_channel'] == 'skip':
            return ctx
        elif channel_name['commands_channel'] is not None and len(channel_name['commands_channel']) > 0:
            channel = bot.get_channel(int(channel_name['commands_channel']))
            return channel
        else:
            return ctx
    else:
        return discord.utils.get(bot.get_all_channels(), guild__name=ctx.guild.name, name=string)


@bot.event
async def on_member_update(before, after):
    authorized = False
    user_id = before.id
    guild_id = before.guild.id
    roles = after.roles
    query = f"SELECT * FROM discords_list WHERE `discord_id` = %s"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query, [guild_id])
            get_guild_data = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return
    if get_guild_data is not None and get_guild_data['view_stockpiles_role'] is not None and len(get_guild_data['view_stockpiles_role']) > 0:

        for role in roles:
            if role.id == int(get_guild_data['view_stockpiles_role']):
                authorized = True

        query = f"SELECT idhash FROM stockpiles_list WHERE `discordid` = %s"
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query, [get_guild_data['id']])
                get_stockpiles = cursor.fetchall()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return

        for h in get_stockpiles:
            query = f"SELECT authorized, name, id FROM stock_users WHERE `user_id` = %s and `stockpile_id` = %s"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query, [user_id, h['idhash']])
                    get_stockpile_user_data = cursor.fetchone()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            if get_stockpile_user_data is not None:
                if authorized and get_stockpile_user_data['authorized'] != 1:
                    query = "UPDATE stock_users SET `authorized` = %s, `name` = %s WHERE `id` = %s"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query, [1, after.display_name, get_stockpile_user_data['id']])
                            connection.commit()
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
                elif not authorized and get_stockpile_user_data['authorized'] != 0:
                    query = "UPDATE stock_users SET `authorized` = %s, `name` = %s WHERE `id` = %s"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query, [0, after.display_name, get_stockpile_user_data['id']])
                            connection.commit()
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
                elif after.name != get_stockpile_user_data['name']:
                    query = "UPDATE stock_users SET `name` = %s WHERE `id` = %s"
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query, [after.display_name, get_stockpile_user_data['id']])
                            connection.commit()
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
            else:
                # -- Insert user data into stock_users -- #
                query = f"INSERT INTO stock_users (stockpile_id, authorized, user_id, name) VALUES (%s, %s, %s, %s)"
                try:
                    connection.ping()
                    with connection.cursor() as cursor:
                        cursor.execute(query, [h['idhash'], authorized, user_id, after.display_name])
                        connection.commit()
                except pymysql.err.ProgrammingError as except_detail:
                    print(except_detail)
                    return


@bot.event
async def on_ready():
    global running

    if running:
        return
    else:
        running = True
    if not debug:
        await check_updates()
    print("Online")

    global active_stockpiles
    stockpiler_checkchanges = time.time()
    run_stockpiler_every_seconds = 5

    stockpiler_5min_running_average = 0
    stockpiler_5min_running_average_timer = time.time()

    stockpiler_60min_running_average_timer = time.time()

    current_log_id = 0

    query = "SELECT * FROM stockpiles_list where active = '1'"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            log_id = cursor.fetchall()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return
    for x in log_id:
        if x['last_log_id'] > current_log_id:
            current_log_id = x['last_log_id']
    status = ""
    while True:
        if (time.time() - stockpiler_checkchanges) >= run_stockpiler_every_seconds:
            stockpiler_checkchanges = time.time()
            startTime = time.time()

            if (time.time() - stockpiler_5min_running_average_timer) >= 300:
                print(
                    f"Stockpiler 5 minute running average: {round(stockpiler_5min_running_average / (300 / run_stockpiler_every_seconds), 4)} seconds.")
                stockpiler_5min_running_average = 0
                stockpiler_5min_running_average_timer = time.time()
            if (time.time() - stockpiler_60min_running_average_timer) >= 3600:  # checks if war ended every hour
                check_war_change()


            active_stockpiles = len(log_id)
            if status != "!Stock Watching " + str(active_stockpiles) + " Stockpiles":
                status = "!Stock Watching " + str(active_stockpiles) + " Stockpiles"
                activity = discord.Game(name=status)
                await bot.change_presence(status=None, activity=activity)
            query = f"SELECT * FROM stock_log WHERE `id` > '{current_log_id}'"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    get_log_details = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            query = "SELECT * FROM stockpiles_list where active = '1'"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    log_id = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            if len(log_id) > 0:  # If there is stockpiles in the list for discord to report on
                #await check_expiring_stockpiles(get_stockpiles_needed) # NEEDS FIXED
                if get_log_details is not None:
                    for k in get_log_details:
                        stock = {}
                        for stock_temp in log_id:
                            #print(f"{stock_temp['idhash']} != {k['stockpile_id']}")
                            if stock_temp['idhash'] == k['stockpile_id']:
                                stock = stock_temp

                        if len(stock) > 0:
                            string = ""
                            logid = "0"
                            if k['log_type'] != "change" and k['log_type'] != "correction":
                                continue
                            await apply_messages(stock['discordid'], k, stock['name'],
                                                 stock['idhash'])

                            logid = k['id']
                            print("New Logs, at ID: " + str(logid))
                            current_log_id = logid

                            query = "UPDATE stockpiles_list SET last_log_id = " + str(logid) + " WHERE id = " + str(
                                stock['id'])
                            try:
                                connection.ping()
                                with connection.cursor() as cursor:
                                    cursor.execute(query)
                                    connection.commit()
                            except pymysql.err.ProgrammingError as except_detail:
                                print(except_detail)
                                return

            query = "SELECT * FROM recent_discord_messages WHERE status = 66"  # aka messages that need deleted.
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    delete_messages = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            if delete_messages is not None and len(delete_messages) > 0:
                for x in delete_messages:
                    channel = discord.utils.get(bot.get_all_channels(), guild__name=x['guild_name'],
                                                name=x['channel_name'])

                    msg = await channel.fetch_message(x["message_id"])

                    await msg.delete()

                    query = "DELETE FROM recent_discord_messages WHERE id = %s"
                    data = (x['id'])
                    try:
                        connection.ping()
                        with connection.cursor() as cursor:
                            cursor.execute(query, data)
                            connection.commit()
                    except pymysql.err.ProgrammingError as except_detail:
                        print(except_detail)
                        return
            executionTime = (time.time() - startTime)
            if executionTime > (run_stockpiler_every_seconds - 1):
                print(f"WARNING!! Stockpiler took {executionTime} seconds to run! Thats approching the rerun "
                      f"function time of {run_stockpiler_every_seconds} seconds!")
            else:
                pass
                # print(f"Run time {executionTime}")
            stockpiler_5min_running_average += executionTime
        else:
            await asyncio.sleep(0.1)



@bot.event
async def on_guild_join(guild):
    print(f"Added to discord {guild.name}")
    for channel in guild.text_channels:
        if channel.permissions_for(guild.me).send_messages:
            query = f"SELECT * FROM discords_list WHERE `discord_name` = '{guild.name}'"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query)
                    get_stockpile_old = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            if len(get_stockpile_old) == 0:
                await channel.send('Hello! Use the command **!stock setup** to get started with setting up this bot!')
            else:
                await channel.send(
                    'Hello! Looks like im already setup for this disord. Use the command **!stock setup** to see configurations.')
        break


async def check_expiring_stockpiles(get_stockpiles_needed):
    expire_debug = True  # Shows additional output for stockpiles named 'Test '.. if editing a _dev file

    global stockalarm_lastran
    check = (time.time() - stockalarm_lastran)
    if check < 12:
        return
    else:
        stockalarm_lastran = time.time()
    query = "SELECT * FROM stock_alarms_config"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            alerts_conf = cursor.fetchall()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return

    query = "SELECT * FROM stock_alarms"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            alerts = cursor.fetchall()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return

    query = "SELECT internal_name, last_addition, id, status, name, active FROM stockpiles_list"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            get_stockpile_old = cursor.fetchall()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return
    if get_stockpile_old is None or len(get_stockpile_old) == 0:
        return

    for x in get_stockpile_old:

        if x['last_addition'] is not None:
            my_timez = pytz.timezone(mytimezone)
            server_timez = pytz.timezone(server_time)
            now = datetime.now(my_timez)
            timeleft = now - server_timez.localize(x['last_addition'])

            minutes_left = math.floor(timeleft.total_seconds() / 60)
            minutes_left = 2880 - minutes_left  # 2 days

            for h in get_stockpiles_needed:
                if x["internal_name"] == h["stockpile_id"]:
                    alert_found = False
                    if expire_debug and debug and "test " in x['name'].lower():
                        print(
                            f"{math.floor(minutes_left / 60)} hours {(minutes_left - (math.floor(minutes_left / 60) * 60))} minutes on {x['name']} or {minutes_left} minutes")
                    discordID = int(h['discords'])
                    if alerts_conf is not None:
                        for g in alerts_conf:
                            if g['stockpile_id'] is not None and len(g['stockpile_id']) > 0 and g['stockpile_id'] != h[
                                "stockpile_id"]:
                                continue
                            if g['discord'] == discordID:
                                alert_found = True
                                if minutes_left <= math.floor(
                                        (timedelta(minutes=g['minutes_left']).total_seconds() / 60)):
                                    alert_data = []
                                    if minutes_left < -1440:
                                        continue
                                    for y in alerts:
                                        if y["config_id"] != g["id"]:
                                            continue
                                        else:
                                            alert_data = y
                                    if len(alert_data) != 0:
                                        if alert_data['pingsent'] == 1:
                                            continue
                                    if minutes_left > 60:
                                        time_string = f"{math.floor((minutes_left / 60))} hours {(minutes_left - (math.floor(minutes_left / 60) * 60))} minutes"
                                    else:
                                        time_string = f"{minutes_left} minutes"
                                    message = f"***{x['name']}*** will expire in {time_string}, Add something to it or make a correction to reset this timer in stockpiler!"

                                    query = "SELECT discord_name, channel_name FROM discords_list WHERE id = " + str(
                                        discordID)
                                    try:
                                        connection.ping()
                                        with connection.cursor() as cursor:
                                            cursor.execute(query)
                                            discord_details = cursor.fetchone()
                                    except pymysql.err.ProgrammingError as except_detail:
                                        print(except_detail)
                                        return
                                    if discord_details is not None and len(discord_details) > 0:
                                        if debug and discord_details['discord_name'] != dev_guild_name:
                                            return
                                        channel = discord.utils.get(bot.get_all_channels(),
                                                                    guild__name=discord_details['discord_name'],
                                                                    name=discord_details['channel_name'])
                                        alert_role = g['roles_to_ping']
                                        if alert_role.lower() == 'here' or alert_role.lower() == 'everyone':
                                            message = f'@{alert_role.lower()}\n{message}'
                                        else:
                                            role = ""
                                            for t in channel.guild.roles:
                                                if role == "" and alert_role == t.name:
                                                    role = t
                                            if role != "":
                                                message = f'{role.mention}\n{message}'
                                            else:
                                                message = f"***{alert_role}*** is not a valid role for this discord. " \
                                                          f"Did you mistype the name of the role on Stockpiler? " \
                                                          f"**!stock setup alert list** Error Code: crUsyJwl\n\n{message}"
                                        try:
                                            await channel.send(message)
                                        except discord.errors.Forbidden:
                                            print(
                                                f"Tried to send meesage about expiring stock but 403 Forbidden: {message} Error code: PYYaTUQi")
                                            print(discord_details)
                                        query = f"INSERT INTO stock_alarms (pingsent, stockpile_id, config_id) VALUES (%s, %s, %s)"
                                        try:
                                            connection.ping()
                                            with connection.cursor() as cursor:
                                                cursor.execute(query, [1, x['internal_name'], g['id']])
                                                connection.commit()
                                        except pymysql.err.ProgrammingError as except_detail:
                                            print(except_detail)
                                            return
                                else:
                                    alert_data = []
                                    for y in alerts:
                                        if y["config_id"] != g["id"]:
                                            continue
                                        else:
                                            alert_data = y
                                    if debug:
                                        print(g)
                                        print(alert_data)
                                    if len(alert_data) > 0 and alert_data['pingsent'] == 1 and alert_data[
                                        'stockpile_id'] == x['internal_name']:
                                        query = "DELETE FROM stock_alarms WHERE id = %s"
                                        try:
                                            connection.ping()
                                            with connection.cursor() as cursor:
                                                cursor.execute(query, [alert_data['id']])
                                                connection.commit()
                                        except pymysql.err.ProgrammingError as except_detail:
                                            print(except_detail)
                                            return
                                        continue
                    if not alert_found and minutes_left <= 120:
                        if x['active'] != '2':
                            message = f"***{x['name']}*** will expire in less then 2 hours! Add something to it or make a correction to reset this timer in stockpiler!\n" \
                                      f"This is a friendly non-pinging notice. You can setup pingable alerts for expiring stockpiles with **!stock setup alert**"
                            active_status = "2"
                        else:
                            message = f"***{x['name']}*** has expired and will no longer show up on **!stock list** commands\n" \
                                      f"Add something to it to reset the timer."
                            active_status = "5"

                        query = "SELECT discord_name, channel_name, alert_role FROM discords_list WHERE id = " + str(
                            discordID)
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query)
                                discord_details = cursor.fetchone()
                        except pymysql.err.ProgrammingError as except_detail:
                            print(except_detail)
                            return

                        query = f"UPDATE stockpiles_list SET `active` = %s WHERE `id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, [active_status, x['id']])
                                connection.commit()
                        except pymysql.err.ProgrammingError as except_detail:
                            print(except_detail)
                            return
                        if discord_details is not None and len(discord_details) > 0:
                            print(discord_details)
                            channel = discord.utils.get(bot.get_all_channels(),
                                                        guild__name=discord_details['discord_name'],
                                                        name=discord_details['channel_name'])
                            print(channel)
                            print(message)
                            try:
                                await channel.send(message)
                            except discord.errors.Forbidden:
                                print(
                                    f"Tried to send meesage about expiring stock but 403 Forbidden: {message} Error code: qoFjHhJp")
                                print(discord_details)

                    if x['active'] == '2' and minutes_left > 120:
                        query = f"UPDATE stockpiles_list SET `active` = %s WHERE `id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, ["1", x['id']])
                                connection.commit()
                        except pymysql.err.ProgrammingError as except_detail:
                            print(except_detail)
                            return


async def check_new_connections():
    # channel = discord.utils.get(bot.get_all_channels(), guild__name='Stockpiler', name='bot_channel')
    for guild in bot.guilds:
        print(f'In {guild.id}')
        async for member in guild.fetch_members():
            print(f"{member.name} -> {member.id}")
            print(member.roles)
            print("")


async def apply_messages(idnumber, get_log_details, stockpile_name, stockpile_id):
    query = "SELECT * FROM discords_list WHERE id = " + str(idnumber)

    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            discord_details = cursor.fetchone()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return

    query = f"SELECT * FROM stock_stock WHERE `stockpile_id` = '{stockpile_id}'"
    try:
        connection.ping()
        with connection.cursor() as cursor:
            cursor.execute(query)
            stock_data = cursor.fetchall()
    except pymysql.err.ProgrammingError as except_detail:
        print(except_detail)
        return

    if len(discord_details) > 0:
        temp_array = discord_details['channel_name'].split(",")
        if len(temp_array) == 0:
            temp_array = [discord_details['channel_name']]

        for channel_id in temp_array:
            channel = bot.get_channel(int(channel_id))

            stockimg = generate_stock_image(get_log_details, stock_data, stockpile_name)

            if get_log_details['discord_text'] is not None and len(get_log_details['discord_text']) > 0:
                message = get_log_details['discord_text']
            else:
                message = get_log_details['notes']
            if "has joined this stockpile." in message and get_log_details['id'] != 2:
                return
            string = "**" + stockpile_name + "**: "  # + cleanhtml(message)

            # embed = discord.Embed(color=0xffffff)
            # embed.set_image(url="")

            if string != "":
                await new_log(string, channel, stockpile_id, get_log_details['user_id'], get_log_details['id'],
                              discord_details['discord_name'],
                              channel_id, stockimg)
    return


def cleanhtml(raw_html):
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext


async def new_log(message, channel, stockpile_id, user, logid, guild_name, channel_name, stockimg):
    if debug and guild_name != dev_guild_name:
        return
    try:
        if stockimg is not None:
            message_details = await channel.send(file=discord.File(stockimg))
        else:
            message_details = await channel.send(message)

        query = f"INSERT INTO recent_discord_messages (message_id, stockpile, user, logid, guild_name, channel_name) VALUES (%s, %s, %s, %s, %s, %s)"
        data = (message_details.id, stockpile_id, user, logid, guild_name, channel_name)
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query, data)
                connection.commit()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return

    except:
        print("Error sending message")
        return


def startdiscord(token):
    bot.run(token)


def excelsheet_to_imgs():
    workbook = load_workbook(filename="Foxhole_items.xlsx", data_only=True)
    sheet = workbook.active
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if row[0] is not None and str(len(row[0])) != "":
            try:
                f = open(f"./imgs/icons/items/{row[4]}.png")
                print(f"Already exists {row[4]}.png")
            except:
                try:
                    f = open(f"./imgs/Public Art and Icons/Icons/Item Icons/{row[2]}")
                    try:
                        copyfile(f"./imgs/Public Art and Icons/Icons/Item Icons/{row[2]}",
                                 f"./imgs/icons/items/{row[4]}.png")
                    except:
                        print(f"Unable to copy row {row[0]} - {row[2]} -> {row[4]}.png")
                except:
                    print(f"Unable to locate image for {row[0]} - {row[2]}")


@bot.command(name='update_member_list_876')
async def check_updates(ctx=[], stockpileid=""):
    # TODO Refine this whole system so net over single member for every stockpile is in the database....
    if len(stockpileid) == 0:
        guilds = bot.guilds
    else:
        guilds = [ctx.guild]
    print(f"Checking members for guilds: {len(guilds)} guilds to check.")
    counter = 0
    for g in guilds:
        counter += 1
        print(f"{counter}/{len(guilds)} Checking member list for Guild: {g.name}")
        members = g.members
        query = f"SELECT id, view_stockpiles_role FROM discords_list WHERE `discord_id` = %s"
        try:
            connection.ping()
            with connection.cursor() as cursor:
                cursor.execute(query, [g.id])
                get_guild_data = cursor.fetchone()
        except pymysql.err.ProgrammingError as except_detail:
            print(except_detail)
            return
        if get_guild_data is not None:
            query = f"SELECT id, idhash FROM stockpiles_list WHERE `discordid` = %s"
            try:
                connection.ping()
                with connection.cursor() as cursor:
                    cursor.execute(query, [get_guild_data['id']])
                    get_stockpile_data = cursor.fetchall()
            except pymysql.err.ProgrammingError as except_detail:
                print(except_detail)
                return
            if get_stockpile_data is not None:
                for s in get_stockpile_data:
                    for m in members:
                        authorized = 0
                        if get_guild_data is not None and get_guild_data['view_stockpiles_role'] is not None and len(
                                get_guild_data['view_stockpiles_role']) > 0:

                            for role in m.roles:
                                if role.id == int(get_guild_data['view_stockpiles_role']):
                                    authorized = 1

                        query = f"SELECT id,authorized,name FROM stock_users WHERE `stockpile_id` = %s and `user_id` = %s"
                        try:
                            connection.ping()
                            with connection.cursor() as cursor:
                                cursor.execute(query, [s['idhash'], m.id])
                                get_stockpile_members = cursor.fetchone()
                        except pymysql.err.ProgrammingError as except_detail:
                            print(except_detail)
                            return
                        if authorized == 1:
                            if get_stockpile_members is None or len(get_stockpile_members) == 0:
                                # -- Insert user data into stock_users -- #
                                query = f"INSERT INTO stock_users (stockpile_id, authorized, user_id, name) VALUES (%s, %s, %s, %s)"
                                try:
                                    connection.ping()
                                    with connection.cursor() as cursor:
                                        cursor.execute(query, [s['idhash'], authorized, m.id, m.display_name])
                                        connection.commit()
                                except pymysql.err.ProgrammingError as except_detail:
                                    print(except_detail)
                                    return
                            elif get_stockpile_members['authorized'] == 0:
                                query = "UPDATE stock_users SET `authorized` = %s, `name` = %s WHERE `id` = %s"
                                try:
                                    connection.ping()
                                    with connection.cursor() as cursor:
                                        cursor.execute(query, [0, after.display_name, get_stockpile_members['id']])
                                        connection.commit()
                                except pymysql.err.ProgrammingError as except_detail:
                                    print(except_detail)
                                    return
                            elif get_stockpile_members['name'] != m.nick:
                                query = "UPDATE stock_users SET `name` = %s WHERE `id` = %s"
                                try:
                                    connection.ping()
                                    with connection.cursor() as cursor:
                                        cursor.execute(query, [m.display_name, get_stockpile_members['id']])
                                        connection.commit()
                                except pymysql.err.ProgrammingError as except_detail:
                                    print(except_detail)
                                    return
                        elif get_stockpile_members is not None and len(get_stockpile_members) > 0 and get_stockpile_members['authorized'] == 1:
                            query = "UPDATE stock_users SET `authorized` = %s, `name` = %s WHERE `id` = %s"
                            try:
                                connection.ping()
                                with connection.cursor() as cursor:
                                    cursor.execute(query, [0, after.display_name, get_stockpile_members['id']])
                                    connection.commit()
                            except pymysql.err.ProgrammingError as except_detail:
                                print(except_detail)
                                return
                        elif get_stockpile_members is not None and len(get_stockpile_members) > 0 and get_stockpile_members['name'] != m.nick:
                            query = "UPDATE stock_users SET `name` = %s WHERE `id` = %s"
                            try:
                                connection.ping()
                                with connection.cursor() as cursor:
                                    cursor.execute(query, [m.display_name, get_stockpile_members['id']])
                                    connection.commit()
                            except pymysql.err.ProgrammingError as except_detail:
                                print(except_detail)
                                return

        print(f"--Complete checking member list for {g.name}")


@bot.command(name='test')
async def test(ctx):
   for x in item_list:
       print(item_list[x])


if __name__ == "__main__":
    key = 'discord_bot_token'
    if "_dev" in os.path.basename(__file__):
        debug = True
        key = 'discord_bot_token'
    p = Process(target=startdiscord, args=(key,))
    p.start()
    # test()